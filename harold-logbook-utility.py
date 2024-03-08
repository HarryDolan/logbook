#!/usr/bin/env python

import pdb
import sys
import argparse
import csv
import numpy as np
import matplotlib.pyplot as plt
import copy
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.styles import PatternFill
from enum import IntEnum, auto

###############################################################################
#
def GetCommandLineArgs ():
    parser = argparse.ArgumentParser(description="Only works on Harold's logbooks.")
    parser.add_argument('--inFF',    help="Input file in ForeFlight's csv format")
    parser.add_argument('--outXLS',  help="Output file in Excel format")
    parser.add_argument('--print',   help="Print output", action='store_true')
    parser.add_argument('--summary', help="Print summary", action='store_true')
    parser.add_argument('--actable', help="Print aircraft table", action='store_true')
    parser.add_argument('--fromDate',help="Skip dates before this")
    parser.add_argument('--toDate',  help="Skip dates after this")
    parser.add_argument('--toPage',  help="Skip pages after this", type=int)
    args = parser.parse_args()
    return args

###############################################################################
#

fieldsFF = ['Date', 'AircraftID', 'From', 'To', 'Route', 'TimeOut',
            'TimeOff', 'TimeOn', 'TimeIn', 'OnDuty', 'OffDuty', 'TotalTime',
            'PIC', 'SIC', 'Night', 'Solo', 'CrossCountry', 'NVG', 'NVG Ops',
            'Distance', 'DayTakeoffs', 'DayLandingsFullStop',
            'NightTakeoffs', 'NightLandingsFullStop', 'AllLandings',
            'ActualInstrument', 'SimulatedInstrument', 'HobbsStart',
            'HobbsEnd', 'TachStart', 'TachEnd', 'Holds', 'Approach1',
            'Approach2', 'Approach3', 'Approach4', 'Approach5', 'Approach6',
            'DualGiven', 'DualReceived', 'SimulatedFlight',
            'GroundTraining', 'InstructorName', 'InstructorComments',
            'Person1', 'Person2', 'Person3', 'Person4', 'Person5',
            'Person6', 'FlightReview', 'Checkride', 'IPC',
            'NVG Proficiency', 'FAA6158', 'PilotComments']

fieldsAll = copy.copy (fieldsFF)

fieldsAll.append ('LogbookPage')
fieldsAll.append ('Approaches')
fieldsAll.append ('Make')
fieldsAll.append ('GliderHeli')
fieldsAll.append ('Helicopter')
fieldsAll.append ('SEL')
fieldsAll.append ('MEL')
fieldsAll.append ('Day')
fieldsAll.append ('Simulator')

entryAll = {}
for field in fieldsAll:
    entryAll[field] = ''

fieldsAC = ['AircraftID', 'EquipmentType', 'TypeCode', 'Year', 'Make', 'Model',
            'Category', 'Class', 'GearType', 'EngineType', 'Complex', 'TAA',
            'HighPerformance', 'Pressurized', '', '', '', '', '', '', '', '',
            '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
            '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']



###############################################################################
#
def parseFF (path, after_date):
    '''For files exported by ForeFlight'''
    
    logEntries = []
    acTable = {}

    with open (path, 'r') as csvfile:
        reader = csv.reader (csvfile)
        inHeader  = False
        inAircraft= False
        fieldsAC  = None
        inFlights = False
        section = None
        for cols in reader:
            if cols[0]=='ForeFlight Logbook Import':
                    section = cols[0]
            elif cols[0]=='Aircraft Table':
                    section = cols[0]
            elif cols[0]=='Flights Table':
                    section = cols[0]
            elif section=='ForeFlight Logbook Import':
                continue
            elif section=='Aircraft Table':
                if cols[0]=='AircraftID':
                    fieldsAC = cols
                else:
                    ac = {}
                    for i in range(len(cols)):
                            ac[fieldsAC[i]] = cols[i]
                    acTable[cols[0]] = cols[1:]

            elif section=='Flights Table':
                if cols[0]=='Date':
                    continue
                else:
                    entry = copy.copy (entryAll)
                    if len(cols[-1]) and cols[-1][0]=='"' and cols[-1][-1]=='"':
                        cols[-1] = cols[-1][1:-1]
                    for i in range(len(fieldsFF)):
                        entry[fieldsFF[i]] = cols[i]
                    if entry['Night']:
                        entry['Day'] = str(float(entry['TotalTime']) - float(entry['Night']))
                    else:
                        entry['Day'] = entry['TotalTime']

                    category = acTable[entry['AircraftID']][6]
                    if category == 'airplane_single_engine_land':
                        entry['SEL'] = entry['TotalTime']
                    elif category == 'airplane_multi_engine_land':
                        entry['MEL'] = entry['TotalTime']
                    elif category == 'rotorcraft_helicopter':
                        entry['Helicopter'] = entry['TotalTime']
                    elif category == 'glider':
                        entry['GliderHeli'] = entry['TotalTime']


                    entry['LogbookPage'] = -99

                    if entry['Approach1']!='':     # Yeah, it's another kludge
                        entry['Approaches']=1
                    if entry['Approach2']!='':
                        entry['Approaches']=2
                    if entry['Approach3']!='':
                        entry['Approaches']=3
                    if entry['Approach4']!='':
                        entry['Approaches']=4
                    if entry['Approach5']!='':
                        entry['Approaches']=5
                    if entry['Approach6']!='':
                        entry['Approaches']=6

                    if entry['AircraftID'] in acTable:
                        entry['Make'] = acTable[entry['AircraftID']][1]

                    if not after_date or entry['Date']>after_date:
                        logEntries.append (entry)

    logEntries.reverse()
    return (logEntries, acTable)


###############################################################################
#
def get_xls_flight_columns ():
    white   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
    gray    = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type = "solid")
    columns = [
        {'index':  'Date',                 'type': 'date',  'color': white, 'width':  82/6, 'pwidth': 11, 'ptitle':'Date',     'title':'DATE'},
        {'index':  'Make',                 'type': 'text',  'color': white, 'width':  66/6, 'pwidth': 10, 'ptitle':'Model',    'title':'AIRCRAFT MODEL'},
        {'index':  'AircraftID',           'type': 'text',  'color': white, 'width':  52/6, 'pwidth': 10, 'ptitle':'Ident',    'title':'AIRCRAFT IDENT.'},
        {'index':  'From',                 'type': 'text',  'color': white, 'width':  48/6, 'pwidth':  5, 'ptitle':'From',     'title':'FROM'},
        {'index':  'To',                   'type': 'text',  'color': white, 'width':  48/6, 'pwidth':  5, 'ptitle':'To',       'title':'TO'},
        {'index':  'Route',                'type': 'text',  'color': white, 'width':  76/6, 'pwidth': 10, 'ptitle':'Route',    'title':'ROUTE'},
        {'index':  'PilotComments' ,       'type': 'text',  'color': white, 'width': 196/6, 'pwidth': 40, 'ptitle':'Comments', 'title':'COMMENTS'},
        {'index':  'Approaches',           'type': 'int',   'color': gray,  'width':  30/6, 'pwidth':  5, 'ptitle':'Appc',     'title':'TOT APPC'},
        {'index':  'AllLandings',          'type': 'int',   'color': gray,  'width':  30/6, 'pwidth':  5, 'ptitle':'Ldgs',     'title':'TOT LDGS'},
        {'index':  'NightLandingsFullStop','type': 'int',   'color': gray,  'width':  30/6, 'pwidth':  5, 'ptitle':'Night',    'title':'NIGHT LDGS'},
        {'index':  'SEL',                  'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'SEL',      'title':'SEL'},
        {'index':  'MEL',                  'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'MEL',      'title':'MEL'},
        {'index':  'GliderHeli',           'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'Gldr.',    'title':'GLDR'},
        {'index':  'Helicopter',           'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'Heli.',    'title':'HELI.'},
        {'index':  'Simulator',            'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'Simu.',    'title':'SIMU.'},
        {'index':  'CrossCountry',         'type': 'float', 'color': gray,  'width':  42/6, 'pwidth':  8, 'ptitle':'X/C',      'title':'X/C'},
        {'index':  'Day',                  'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'Day',      'title':'DAY'},
        {'index':  'Night',                'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'Night',    'title':'NIGHT'},
        {'index':  'ActualInstrument',     'type': 'float', 'color': gray,  'width':  42/6, 'pwidth':  8, 'ptitle':'Act.Ins.', 'title':'ACTUAL INSTR.'},
        {'index':  'SimulatedInstrument',  'type': 'float', 'color': gray,  'width':  42/6, 'pwidth':  8, 'ptitle':'Hooded',   'title':'HOODED'},
        {'index':  'PIC',                  'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'PIC',      'title':'PIC'},
        {'index':  'SIC',                  'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'SIC',      'title':'SIC'},
        {'index':  'DualGiven',            'type': 'float', 'color': gray,  'width':  42/6, 'pwidth':  8, 'ptitle':'Dual G',   'title':'DUAL GIVEN'},
        {'index':  'DualReceived',         'type': 'float', 'color': gray,  'width':  42/6, 'pwidth':  8, 'ptitle':'Dual R',   'title':'DUAL REC\'D'},
        {'index':  'TotalTime',            'type': 'float', 'color': white, 'width':  42/6, 'pwidth':  8, 'ptitle':'Total',    'title':'TOTAL'}
    ]

    return columns

###############################################################################
#
def printLogEntries (logEntries, args):
    columns = get_xls_flight_columns()
    COL_NUMBERS_START = 8
    tot_page    = [0]*len(columns)
    tot_forward = [0]*len(columns)
    tot_todate  = [0]*len(columns)
    for c in range(COL_NUMBERS_START-1,len(columns)):
        if columns[c]['type']=='int':
            tot_forward[c] = 0
        elif columns[c]['type']=='float':
            tot_forward[c] = 0.0


    entnum = 0
    for entry in logEntries:
        entnum += 1
        page = int(entnum / 7 + 1)
        if entnum%7==1:
            print ('Page ', page)
            c = 0
            for col in columns:
                c += 1
#                if col['type']==int or col['type']==float:
                if c>=COL_NUMBERS_START:
                    print ('{:>{wid}s}'.format(col['ptitle'],wid=col['pwidth']), end='')
                else:
                    print ('{:<{wid}s}'.format(col['ptitle'],wid=col['pwidth']), end='')
            print ()

        c = 0
        for col in columns:
            val = entry[col['index']]
            typ = col['type']
            pwidth = col['pwidth']
            if typ=='date':
                str = '{:{wid}s}'.format(val,wid=pwidth)
            elif typ=='text':
                if len(val)>pwidth: val = val[0:pwidth]
                str = '{:{wid}s}'.format(val,wid=pwidth)
            elif typ=='int':
                try:
                    str = '{:>{wid}d}'.format(val,wid=pwidth)
                    tot_page[c] += val
                except:
                    str = '{:>{wid}s}'.format(val,wid=pwidth)
                    if val != '':
                        tot_page[c] += int(val)
            elif typ=='float':
                try:
                    str = '{:>{wid}.1f}'.format(val,wid=pwidth)
                    tot_page[c] += val
                except:
                    str = '{:>{wid}s}'.format(val,wid=pwidth)
                    if val != '':
                        tot_page[c] += float(val)
            else:
                str = 'xxxxx'
            print (str, sep='', end='')
            c += 1
        print ()
        if entnum%7==0:
            print (76*' ', 'Page total    ', end='')
            for  c in range(COL_NUMBERS_START-1,len(columns)):
                if columns[c]['type']=='int':
                    str = '{:{wid}d}'.format(tot_page[c],wid=columns[c]['pwidth'])
                elif columns[c]['type']=='float':
                    str = '{:{wid}.1f}'.format(tot_page[c],wid=columns[c]['pwidth'])
                else:
                    str = 'yyyyy'
                print (str, sep='', end='')
            print ()
            print (76*' ', 'Amt. forward  ', end='')
            for  c in range(COL_NUMBERS_START-1,len(columns)):
                if columns[c]['type']=='int':
                    str = '{:{wid}d}'.format(tot_forward[c],wid=columns[c]['pwidth'])
                elif columns[c]['type']=='float':
                    str = '{:{wid}.1f}'.format(tot_forward[c],wid=columns[c]['pwidth'])
                else:
                    str = 'yyyyy'
                print (str, sep='', end='')
            print ()
            print (76*' ', 'Total to date ', end='')
            for  c in range(COL_NUMBERS_START-1,len(columns)):
                tot_todate[c] += tot_page[c]
                if columns[c]['type']=='int':
                    str = '{:{wid}d}'.format(tot_todate[c],wid=columns[c]['pwidth'])
                elif columns[c]['type']=='float':
                    str = '{:{wid}.1f}'.format(tot_todate[c],wid=columns[c]['pwidth'])
                else:
                    str = 'yyyyy'
                print (str, sep='', end='')
                tot_forward[c] = tot_todate[c]
                tot_page[c] = 0
            print ()

        if args.toPage is not None and page>args.toPage: break



###############################################################################
#
def calculateSummaries (logEntries):
    aircraft = {}
    acMake = {}
    airportFrom = {}
    airportTo = {}
    route = {}

    for entry in logEntries:
        if entry['AircraftID'] not in aircraft:
            aircraft[entry['AircraftID']] = 1
        else:
            aircraft[entry['AircraftID']] += 1

        try:
            tottime = float(entry['TotalTime'])
        except:
            tottime = 0.0

        if entry['Make']:
            if entry['Make'] not in acMake:
                acMake[entry['Make']] = tottime
            else:
                acMake[entry['Make']] += tottime

        if entry['From'] not in airportFrom:
            airportFrom[entry['From']] = 1
        else:
            airportFrom[entry['From']] += 1

        if entry['To'] not in airportTo:
            airportTo[entry['To']] = 1
        else:
            airportTo[entry['To']] += 1

        if entry['Route'] not in route:
            route[entry['Route']] = 1
        else:
            route[entry['Route']] += 1

    totals = {'PIC': 0.0,
              'SIC': 0.0,
              'Night': 0.0,
              'Solo': 0.0,
              'CrossCountry': 0.0,
              'DualGiven': 0.0,
              'DualReceived': 0.0,
              'NightLandingsFullStop': 0.0,
              'AllLandings': 0.0,
              'ActualInstrument': 0.0,
              'SimulatedInstrument': 0.0,
              'TotalTime': 0.0}
    
    for entry in logEntries:
        for time in totals.keys():
            if entry[time]:
                totals[time] += float(entry[time])

    return totals, acMake

###############################################################################
#
def printSummaries (logEntries):

    totals, acMake = calculateSummaries (logEntries)

    total = 0.
    for ma in sorted(acMake.keys()):
        round = float( int(acMake[ma] * 10 + .001)) / 10.
        total += round
        print ('{:<15s} {:10.1f}'.format(ma, round))
    print ('{:<15s} {:10.1f}'.format('Total', total))
    print (26*'-')
    for time in totals.keys():
        print ('{:<22s} {:10.1f}'.format(time, totals[time]))
    return

###############################################################################
#
def printAircraftTable (acTable):

    for key in acTable.keys():
        print (key, acTable[key])

    return

###############################################################################
#
def outXLS (acTable, logEntries, fileXLS):
    wb = Workbook()
    wb.iso_dates = True
    ws1 = wb.active
    ws1.title = "Flights"
    ws2 = wb.create_sheet(title="Aircraft")
    ws3 = wb.create_sheet(title="Summary")

    class Row (IntEnum):
        PAGE = 1
        HEAD1 = 2
        ENTRY = 3
        ENTRY_LAST = ENTRY + 6
        PAGE_TOT = ENTRY_LAST + 1
        FWD = PAGE_TOT + 1
        TOTAL = FWD + 1
        SPACER = TOTAL + 1

    yellow_bgd  = PatternFill(start_color="FFFFA5", end_color="FFFFA5", fill_type = "solid")
    gray_bgd    = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type = "solid")
    black_bgd   = PatternFill(start_color="000000", end_color="000000", fill_type = "solid")
    white   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
    gray    = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type = "solid")

    font = Font (name='Arial', size=8, bold=False, italic=False, vertAlign=None, underline='none', strike=False)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    columns = get_xls_flight_columns()

    for c in range(len(columns)):
        if c<26:
            let1 = ''
        else:
            let1 = chr(ord('A')+int(c/26)-1)
        let2 = chr(ord('A')+c%26)
        columns[c]['letter'] = let1 + let2

    NROWS = Row.SPACER
    NCOLS = len(columns)
    COL_NUMBERS_START = 8

    for c in range(len(columns)):
        ws1.column_dimensions[columns[c]['letter']].width = columns[c]['width']

    entnum = 0
    for entry in logEntries:

        page = int(entnum / 7 + 1)
        row1 = int((page-1) * NROWS)
        entrow = row1 + entnum % 7 + Row.ENTRY

        if entnum%7 == 0:

            print ('page={:d} entnum={:d} row1={} entrow={:d}'.format(page,entnum,row1,entrow))

            # Create all cells in this page:
            for r in range (row1+1, row1+1+NROWS):
                for c in range (1, NCOLS+1):
                    ws1.cell(row=r,column=c,value='')
                    ws1.cell(row=r,column=c).border = thin_border
                    ws1.cell(row=r,column=c).font = font
                    if c==7:
                        ws1.cell(row=r,column=c).alignment = Alignment(wrapText=True,horizontal='left',vertical='center')
                    else:
                        ws1.cell(row=r,column=c).alignment = Alignment(wrapText=True,horizontal='center',vertical='center')
            # Page row:
            ws1.row_dimensions[row1+Row.PAGE].height = 20
            ws1.cell (row=row1+Row.PAGE, column=1, value='Page ' + str(page))
            for c in range (2, NCOLS+1):
                ws1.cell(row=row1+Row.PAGE, column=c).fill = gray_bgd
            ws1.merge_cells(start_row=row1+Row.PAGE, start_column=2, end_row=row1+Row.PAGE, end_column=NCOLS)

            # Header row:
            ws1.row_dimensions[row1+Row.HEAD1].height = 30
            for c in range (1, NCOLS+1):
                ws1.cell(row=row1+Row.HEAD1,column=c).fill = yellow_bgd
            for c in range(len(columns)):
                ws1.cell (row=row1+Row.HEAD1, column=c+1, value=columns[c]['title'])

            # Entry rows (7):
            for r in range (row1+Row.ENTRY, row1+Row.ENTRY+7):
                ws1.row_dimensions[r].height = 35
                for c in range(NCOLS):
                    ws1.cell(row=r,column=c+1).fill = columns[c]['color']

            # Formulas for Page Totals:
            ws1.row_dimensions[row1+Row.PAGE_TOT].height = 13
            ws1.merge_cells(start_row=row1+Row.PAGE_TOT, start_column=1, end_row=row1+Row.TOTAL, end_column=COL_NUMBERS_START-2)
            ws1.cell (row=row1+Row.PAGE_TOT, column=COL_NUMBERS_START-1, value='PAGE TOTAL')
            for c in range (COL_NUMBERS_START-1, len(columns)+1):
                ws1.cell(row=row1+Row.PAGE_TOT,column=c).fill = yellow_bgd
            for c in range(COL_NUMBERS_START-1,len(columns)):
                ws1.cell (row=row1+Row.PAGE_TOT, column=c+1,
                          value='=SUM(' + columns[c]['letter']
                          + str(row1+Row.ENTRY)
                          + ':' + columns[c]['letter']
                          + str(row1+Row.ENTRY+6)
                          + ')')

            # Formulas for Amounts Forward:
            ws1.row_dimensions[row1+Row.FWD].height = 13
            ws1.cell (row=row1+Row.FWD, column=COL_NUMBERS_START-1, value='AMT. FORWARD')
            for c in range (COL_NUMBERS_START-1, len(columns)+1):
                ws1.cell(row=row1+Row.FWD,column=c).fill = yellow_bgd
            if row1>0:
                for c in range(COL_NUMBERS_START-1,len(columns)):
                    ws1.cell (row=row1+Row.FWD, column=c+1,
                              value='=(' + columns[c]['letter']
                              + str(row1-NROWS+Row.TOTAL)
                              + ')')

            # Formulas for Totals To Date:
            ws1.row_dimensions[row1+Row.TOTAL].height = 13
            ws1.cell (row=row1+Row.TOTAL, column=COL_NUMBERS_START-1, value='TOTAL TO DATE')
            for c in range (COL_NUMBERS_START-1, len(columns)+1):
                ws1.cell(row=row1+Row.TOTAL,column=c).fill = yellow_bgd
            for c in range(COL_NUMBERS_START-1,len(columns)):
                ws1.cell (row=row1+Row.TOTAL, column=c+1,
                          value='=SUM(' + columns[c]['letter']
                          + str(row1+Row.PAGE_TOT)
                          + ':' + columns[c]['letter']
                          + str(row1+Row.FWD)
                          + ')')

            # Spacer row:
            ws1.row_dimensions[row1+Row.SPACER].height = 100
            for c in range (1, NCOLS+1):
                ws1.cell(row=row1+Row.SPACER,column=c).fill = gray_bgd
            ws1.merge_cells(start_row=row1+Row.SPACER, start_column=1, end_row=row1+Row.SPACER, end_column=NCOLS)

        # Insert current entry into spreadsheet:
        for c in range(len(columns)):
            index = columns[c]['index']
            if index!='':
                if index in entry:
                    if entry[index]:
                        if entry[index]==0 or entry[index]==0.0 or entry[index]=='0' or entry[index]=='0.0' or entry[index]=='0.00':  #KLUDGE ALERT!
                            entry[index] = ''
                        if type(entry[index])==str:
                            ent = entry[index].strip()
                        else:
                            ent = entry[index]
                        if ent=='':
                            ws1.cell (row=entrow, column=c+1, value=ent)
                        elif columns[c]['type']=='int':
                            ws1.cell (row=entrow, column=c+1, value=int(ent))
                        elif columns[c]['type']=='float':
                            ws1.cell (row=entrow, column=c+1, value=float(ent))
                            ws1.cell (row=entrow, column=c+1).number_format = '##,##0.0'
                        else:
                            ws1.cell (row=entrow, column=c+1, value=ent)

        entnum += 1

    headings = ['AircraftID', 'EquipmentType', 'TypeCode', 'Year', 'Make',
                'Model', 'Category', 'Class', 'GearType', 'EngineType',
                'Complex', 'TAA', 'HighPerformance', 'Pressurized']

    c = 0
    for h in headings:
        c += 1
        ws2.cell (row=1, column=c, value=h)
        letter = chr(ord('A')+c%26)
        if letter=='C':
            width = 58/6
        elif letter=='E':
            width = 72/6
        elif letter=='F':
            width = 64/6
        elif letter=='H':
            width = 138/6
        elif letter=='I':
            width = 96/6
        elif letter=='J':
            width = 72/6
        else:
            width = 52/6
        ws2.column_dimensions[columns[c]['letter']].width = width

    row = 1
    for ac in sorted(acTable.keys()):
        row += 1
        ws2.cell (row=row, column=1, value=ac)
        col = 1
        for val in acTable[ac]:
            col += 1
            ws2.cell (row=row, column=col, value=val)
            
    # Summaries:

    totals, acMake = calculateSummaries (logEntries)

    row = 0
    for key in totals.keys():
        row += 1
        ws3.cell (row=row, column=1, value=key)
        ws3.cell (row=row, column=2, value=totals[key])

    total = 0.
    row = 0
    for key in sorted(acMake.keys()):
        row += 1
        ws3.cell (row=row, column=4, value=key)
        ws3.cell (row=row, column=5, value=acMake[key])
        round = float( int(acMake[key] * 10 + .001)) / 10.
        total += round
    ws3.cell (row=row+1, column=4, value='Total')
    ws3.cell (row=row+1, column=5, value=total)
    ws3.column_dimensions['A'].width = 108/6
    ws3.column_dimensions['D'].width = 84/6
    
    wb.save(filename = fileXLS)

###############################################################################
#
def main():
    args = GetCommandLineArgs()
    logEntries = []
    acTable = {}
    if args.inFF:
        if len(logEntries)>0:
            after_date = entries[-1]['Date']   # Ignore FF entries that are also in XLS
        else:
            after_date = None
        entries,aircraft = parseFF (args.inFF, after_date)
        logEntries += entries
        acTable.update(aircraft)

    if args.fromDate:
        ifrom = 0
        for i in range(len(logEntries)):
            if str(logEntries[i]['Date']) < args.fromDate:
                ifrom = i+1
        logEntries = logEntries[ifrom:]
    if args.toDate:
        ito = 0
        for i in range(len(logEntries)):
            if str(logEntries[i]['Date']) > args.toDate:
                ito = i
                break
        logEntries = logEntries[:ito]

    if args.print:
        printLogEntries (logEntries, args)
    if args.actable:
        printAircraftTable (acTable)
    if args.summary:
        printSummaries (logEntries)
    if args.outXLS:
        outXLS (acTable, logEntries, args.outXLS)

###############################################################################
#
if __name__ == "__main__":
    main()
